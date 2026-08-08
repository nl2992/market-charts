#!/usr/bin/env python3
"""
Build Market_Charts.xlsx — one workbook, five charts on a single sheet.

  1  Equities, bonds and commodities rebased to 100
  2  Initial claims against their own seasonal
  3  Continuing claims and the unemployment rate
  4  The 10y through six equity bubbles
  5  The data centre boom and the crowding out

Every chart is a native Excel object pointing at a data sheet in the same
file, so anything on a chart can be traced to the number behind it and the
identifier it was pulled with. Nothing is pasted in as a picture.

Sources are public and unauthenticated: Yahoo Finance for prices, FRED for
claims, the unemployment rate and Treasury yields, and the Census C30
release for construction.

Run:  python3 data-puller/pull_cross_asset.py
      python3 data-puller/pull_duration_case.py
      python3 data-puller/build_public_charts.py
"""

import csv
import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import Reference
from openpyxl.styles import Alignment

import hsbc_xlsx as hx

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")
OUT = os.path.join(HERE, "Market_Charts.xlsx")

AUTHOR = "Nigel Li"
HDR, R0, C0 = 6, 7, 2

# The three asset classes for chart 1, and the series standing in for each.
ASSETS = [
    ("ACWI", "Equities — MSCI ACWI", hx.RED),
    ("UST710", "Bonds — UST 7-10y", hx.NAVY),
    ("GSCI", "Commodities — S&P GSCI", hx.TEAL),
]


def load(name):
    with open(os.path.join(DATA, name)) as fh:
        return list(csv.DictReader(fh))


def val(row, key):
    v = row.get(key, "")
    return None if v in ("", None) else float(v)


def grid(ws, first_label, labels, width=14, first_width=12):
    c = ws.cell(row=HDR, column=1, value=first_label)
    c.fill, c.font = hx.FILL_RED, hx.F_HEAD
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.column_dimensions["A"].width = first_width
    for i, lab in enumerate(labels):
        c = ws.cell(row=HDR, column=C0 + i, value=lab)
        c.fill, c.font = hx.FILL_RED, hx.F_HEAD
        c.alignment = Alignment(horizontal="right", vertical="center",
                                wrap_text=True)
        ws.column_dimensions[hx.get_column_letter(C0 + i)].width = width
    ws.row_dimensions[HDR].height = 30
    ws.freeze_panes = ws.cell(row=R0, column=C0)


def sheet_cross_asset(wb):
    """Prices as values, the rebasing as live formulas so the base date can
    be moved without a rebuild."""
    rows = load("cross_asset_daily.csv")
    ws = wb.create_sheet("CrossAsset")
    hx.head(ws, "Equities, bonds and commodities",
            "Daily close. Columns B-D are the pulled prices; E-G rebase them "
            "to 100 at the first row.", width=8)
    labels = [a[1] for a in ASSETS] + ["%s, rebased" % a[1] for a in ASSETS]
    grid(ws, "Date", labels, width=17)

    n = len(rows)
    for r, row in enumerate(rows, start=R0):
        c = ws.cell(row=r, column=1, value=dt.date.fromisoformat(row["date"]))
        c.number_format, c.font = hx.FMT_DATE, hx.F_BODY
        for i, (key, _lab, _col) in enumerate(ASSETS):
            p = ws.cell(row=r, column=C0 + i, value=val(row, key))
            p.number_format, p.font = hx.FMT_PX, hx.F_BODY
            src = hx.get_column_letter(C0 + i)
            q = ws.cell(row=r, column=C0 + 3 + i,
                        value="=100*%s%d/%s$%d" % (src, r, src, R0))
            q.number_format, q.font = hx.FMT_IDX, hx.F_BODY
    return ws, n


def sheet_claims(wb):
    rows = load("claims_weekly.csv")
    ws = wb.create_sheet("Claims")
    hx.head(ws, "Initial jobless claims",
            "Seasonally adjusted, thousands, by week of year. Series ICSA.",
            width=6)
    years = ["2026", "2025", "2024", "2023"]
    grid(ws, "Week", years + ["2023-25 average"])
    n = len(rows)
    for r, row in enumerate(rows, start=R0):
        ws.cell(row=r, column=1, value=int(row["week"])).font = hx.F_BODY
        for i, y in enumerate(years):
            v = val(row, y)
            c = ws.cell(row=r, column=C0 + i,
                        value=None if v is None else v / 1000.0)
            c.number_format, c.font = "#,##0", hx.F_BODY
        # A flat reference so the eye has something to read 2026 against.
        # Live, so it moves if the data does.
        c = ws.cell(row=r, column=C0 + 4,
                    value="=AVERAGE($C$%d:$E$%d)" % (R0, R0 + n - 1))
        c.number_format, c.font = "#,##0", hx.F_BODY
    return ws, n, years


def sheet_continuing(wb):
    rows = load("claims_urate.csv")
    ws = wb.create_sheet("Continuing")
    hx.head(ws, "Continuing claims and the unemployment rate",
            "Change from November, survey reference week (the week containing "
            "the 12th). Series CCSA and UNRATE.", width=6)
    cols = ["cc_2026", "cc_2024", "ur_2026", "ur_2024"]
    titles = ["Continuing claims, this cycle", "Continuing claims, 2024",
              "Unemployment rate, this cycle", "Unemployment rate, 2024"]
    grid(ws, "Month", titles, width=18)
    for r, row in enumerate(rows, start=R0):
        ws.cell(row=r, column=1, value=row["month"]).font = hx.F_BODY
        for i, k in enumerate(cols):
            c = ws.cell(row=r, column=C0 + i, value=val(row, k))
            c.number_format = "0.0%" if k.startswith("cc") else "0.00"
            c.font = hx.F_BODY
    return ws, len(rows), titles


def sheet_bubbles(wb):
    rows = [r for r in load("bubbles.csv") if int(r["day"]) <= 500]
    ws = wb.create_sheet("Bubbles")
    hx.head(ws, "The 10y through equity bubbles",
            "US 10y Treasury yield, change from the start of each episode, "
            "percentage points. Series DGS10.", width=8)
    cols = ["2025-26 current", "2020-22", "1998-2002 dot-com", "2023-25",
            "1964-66", "1997-98"]
    grid(ws, "Day", cols, width=16)
    for r, row in enumerate(rows, start=R0):
        ws.cell(row=r, column=1, value=int(row["day"])).font = hx.F_BODY
        for i, k in enumerate(cols):
            c = ws.cell(row=r, column=C0 + i, value=val(row, k))
            c.number_format, c.font = "0.00", hx.F_BODY
    return ws, len(rows), cols


def sheet_construction(wb):
    rows = load("construction.csv")
    ws = wb.create_sheet("Construction")
    hx.head(ws, "Private non-residential construction",
            "Change since January 2024, $bn, seasonally adjusted annual rate. "
            "Census C30. Columns E-G repeat the same series in January 2024 "
            "prices.", width=8)
    cols = ["data_center", "total", "ex_data_center",
            "data_center_real", "total_real", "ex_data_center_real"]
    titles = ["Data centres", "Total non-residential", "Ex data centres",
              "Data centres, Jan-24 prices", "Total, Jan-24 prices",
              "Ex data centres, Jan-24 prices"]
    grid(ws, "Month", titles, width=17)
    for r, row in enumerate(rows, start=R0):
        c = ws.cell(row=r, column=1, value=dt.date.fromisoformat(row["date"]))
        c.number_format, c.font = hx.FMT_DATE, hx.F_BODY
        for i, k in enumerate(cols):
            v = ws.cell(row=r, column=C0 + i, value=val(row, k))
            v.number_format = "+#,##0.0;-#,##0.0"
            v.font = hx.F_BODY
    return ws, len(rows), titles


def sheet_betas(wb):
    rows = load("rate_oil_betas.csv")
    ws = wb.create_sheet("RateOilBeta")
    hx.head(ws, "10y rate sensitivity to Brent",
            "Basis points per one per cent move in Brent, from weekly "
            "Friday-to-Friday changes since February 2026.", width=6)
    grid(ws, "Market", ["Sensitivity (bps per % Brent)", "R squared",
                        "Source", "Identifier"], width=20, first_width=14)
    for r, row in enumerate(rows, start=R0):
        ws.cell(row=r, column=1, value=row["label"]).font = hx.F_LBL
        b = ws.cell(row=r, column=C0,
                    value=float(row["beta_bps_per_pct_brent"]))
        b.number_format, b.font = "0.00", hx.F_BODY
        q = ws.cell(row=r, column=C0 + 1, value=float(row["r_squared"]))
        q.number_format, q.font = "0.00", hx.F_BODY
        for j, k in enumerate(["source", "identifier"]):
            c = ws.cell(row=r, column=C0 + 2 + j, value=row[k])
            c.font = hx.F_BODY
    n = len(rows)
    hx.note(ws, "A%d" % (R0 + n + 2),
            "Canada, Switzerland and New Zealand are absent because no free "
            "daily source for their 10y was reachable: the Bank of Canada "
            "benchmark series has stopped returning observations, the SNB "
            "bond cube is monthly and ends July 2025, and the RBNZ blocks "
            "automated requests. They are left out rather than proxied.")
    ws.merge_cells(start_row=R0 + n + 2, start_column=1,
                   end_row=R0 + n + 2, end_column=6)
    ws.row_dimensions[R0 + n + 2].height = 44
    return ws, n


def sheet_charts(wb, refs):
    """All five charts on one sheet, stacked."""
    ws = wb.create_sheet("Charts")
    hx.head(ws, "Charts",
            "Five charts, each pointing at a data sheet in this workbook. "
            "Nothing here is typed in or pasted as a picture.", width=14)
    ws.sheet_view.showGridLines = False

    # Daily data, so the axis is thinned to roughly one label a half-year and
    # the scale starts at 80 rather than nought — a rebased chart that starts
    # at zero wastes four fifths of the plot on space nothing travels through.
    ca, ca_n = refs["cross_asset"]
    ser = [hx.line_series(ca, C0 + 3 + i, R0, ca_n, 1, colour, label,
                          width=24000)
           for i, (_k, label, colour) in enumerate(ASSETS)]
    hx.line_chart(ws, "A5", "Equities, bonds and commodities", ser,
                  y_title="Index, 02-Jan-23 = 100", n_points=ca_n,
                  target_labels=8, x_fmt="mmm-yy", y_fmt="0",
                  y_min=80, y_max=210, height=10.5, width=21.0)

    cl, cl_n, years = refs["claims"]
    ser = [hx.line_series(cl, C0 + i, R0, cl_n, 1, c, y,
                          width=26000 if i == 0 else 14000)
           for i, (y, c) in enumerate(zip(years, [hx.RED, hx.NAVY,
                                                  hx.BLUE_LT, hx.GREY_PALE]))]
    ser.append(hx.line_series(cl, C0 + 4, R0, cl_n, 1, hx.GREY,
                              "2023-25 average", width=9000, dashed=True))
    hx.line_chart(ws, "A28", "Initial claims by week of year", ser,
                  y_title="Thousands", date_axis=False, n_points=cl_n,
                  target_labels=13, y_fmt="#,##0", y_min=170, y_max=270,
                  height=10.5, width=21.0)

    co, co_n, titles = refs["continuing"]
    colours = [hx.RED, hx.GREY_PALE, hx.NAVY, hx.BLUE_LT]
    dashes = [False, False, True, True]
    ser = [hx.line_series(co, C0 + i, R0, co_n, 1, colours[i], titles[i],
                          width=26000 if i == 0 else 16000, dashed=dashes[i])
           for i in range(4)]
    # Claims are a per cent change and the rate a percentage point change, so
    # the rate series need their own scale or the gap looks meaningful.
    hx.combo_chart(ws, "A51", "Continuing claims and the unemployment rate",
                   primary=ser[:2], secondary=ser[2:],
                   y_title="Continuing claims, % change from November",
                   y2_title="Unemployment rate, pp change",
                   n_points=co_n, target_labels=co_n, y_fmt="0%",
                   y2_fmt="0.0", height=10.5, width=21.0)

    bu, bu_n, cols = refs["bubbles"]
    colours = [hx.RED, hx.NAVY, hx.BLUE_MID, hx.TEAL, hx.GREY_PALE, hx.BLUE_LT]
    ser = [hx.line_series(bu, C0 + i, R0, bu_n, 1, colours[i], cols[i],
                          width=26000 if i == 0 else 14000)
           for i in range(len(cols))]
    hx.line_chart(ws, "A74", "10y change from the start of each episode", ser,
                  y_title="Percentage points", date_axis=False, n_points=bu_n,
                  target_labels=11, y_fmt="0.0", y_min=-1.5, y_max=3.5,
                  height=10.5, width=21.0)

    # Data centres go on their own right-hand scale, which is how the
    # sell-side version draws it: the boom is an order of magnitude smaller
    # than the hole it is being contrasted with, and on one axis it flattens.
    cn, cn_n, titles = refs["construction"]
    dc = hx.line_series(cn, C0 + 0, R0, cn_n, 1, hx.RED, titles[0] + " (RHS)",
                        width=26000)
    lhs = [hx.line_series(cn, C0 + 1, R0, cn_n, 1, hx.BLUE_MID, titles[1],
                          width=17000),
           hx.line_series(cn, C0 + 2, R0, cn_n, 1, hx.NAVY, titles[2],
                          width=20000)]
    hx.combo_chart(ws, "A97", "Construction spending, change since January 2024",
                   primary=lhs, secondary=[dc],
                   y_title="Total and ex data centres, $bn",
                   y2_title="Data centres, $bn",
                   n_points=cn_n, target_labels=10, x_fmt="mmm-yy",
                   y_fmt="+#,##0;-#,##0", y2_fmt="+#,##0;-#,##0",
                   height=10.5, width=21.0)

    bt, bt_n = refs["betas"]
    vals = Reference(bt, min_col=C0, min_row=R0, max_row=R0 + bt_n - 1)
    cats = Reference(bt, min_col=1, min_row=R0, max_row=R0 + bt_n - 1)
    hx.bar_chart(ws, "A120", "10y rate sensitivity to Brent", vals, cats,
                 height=10.5, width=21.0, fmt="0.00",
                 series_title="bps per % Brent")
    return ws


def sheet_cover(wb):
    ws = wb.create_sheet("Cover", 0)
    hx.head(ws, "Market charts", "Built from public data", width=6)
    ws.column_dimensions["A"].width = 17
    ws.column_dimensions["B"].width = 84

    rows = [
        ("1  Cross-asset", "MSCI ACWI, UST 7-10y and the S&P GSCI rebased to "
                           "100 at 2 January 2023."),
        ("2  Claims", "US initial jobless claims by week of year, 2023 to 2026."),
        ("3  Continuing", "Continuing claims and the unemployment rate against "
                          "the same month a cycle earlier."),
        ("4  Bubbles", "The US 10y through six equity episodes, measured from "
                       "the start of each."),
        ("5  Construction", "US private non-residential construction: data "
                            "centres against everything else."),
        ("6  Rate-oil beta", "How far each G10 10y moves for a one per cent "
                             "move in Brent, on weekly changes since "
                             "February 2026."),
    ]
    r = 5
    for k, v in rows:
        hx.label(ws, "A%d" % r, k)
        ws["B%d" % r] = v
        ws["B%d" % r].font = hx.F_BODY
        ws["B%d" % r].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    hx.label(ws, "A%d" % r, "Sources", font=hx.F_H2)
    r += 1
    for k, v in [
        ("Yahoo Finance", "ACWI, IEF and ^SPGSCI daily closes. The ETF series "
                          "use the adjusted close, so distributions are "
                          "reinvested; the S&P GSCI is a spot index."),
        ("FRED", "ICSA, CCSA, UNRATE and DGS10."),
        ("US Census Bureau", "C30 value of construction put in place, private, "
                             "seasonally adjusted. The data centre series is "
                             "not carried on FRED and comes from the Census "
                             "file directly."),
        ("FRED", "WPUSI012011, PPI construction materials, used to deflate."),
    ]:
        hx.label(ws, "A%d" % r, k)
        ws["B%d" % r] = v
        ws["B%d" % r].font = hx.F_BODY
        ws["B%d" % r].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 34
        r += 1

    r += 1
    hx.label(ws, "A%d" % r, "Notes", font=hx.F_H2)
    r += 1
    for line in [
        "Rebasing divides each series by its own first close and multiplies "
        "by 100, so every line starts together and the chart reads as "
        "relative performance.",
        "The cross-asset spine is every weekday, with the previous close "
        "carried through a market holiday.",
        "The dashed line on chart 2 is the 2023-25 average across all weeks, "
        "computed live on the Claims sheet. It is a reference for reading "
        "2026 against, not a threshold with any meaning of its own.",
        "Chart 4 measures each episode from its own start date. Yields rose "
        "in four of the six and fell in two. The current episode runs from "
        "26 June 2025 and is not complete.",
        "Chart 5 is nominal, which is how this is usually drawn. Columns E-G "
        "of the Construction sheet repeat it in January 2024 prices. Data "
        "centres sit on the right-hand scale.",
        "Chart 6 is a plain least-squares slope of the weekly change in each "
        "10y, in basis points, on the weekly per cent change in Brent, "
        "Friday to Friday. The estimates move materially with the sampling "
        "day - Britain ranges from 0.83 to 1.11 across Monday, Wednesday and "
        "Friday sampling - so read the ranking, not the level. R squared per "
        "market is on the RateOilBeta sheet.",
    ]:
        r = hx.bullet(ws, r, line)
        ws.merge_cells(start_row=r - 1, start_column=1, end_row=r - 1,
                       end_column=6)
        ws.row_dimensions[r - 1].height = 30

    r += 1
    hx.note(ws, "A%d" % r,
            "Built %s. Rebuilt from primary sources rather than reproduced "
            "from third-party research."
            % dt.date.today().strftime("%d %B %Y"))
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    refs = {}
    refs["cross_asset"] = sheet_cross_asset(wb)
    refs["claims"] = sheet_claims(wb)
    refs["continuing"] = sheet_continuing(wb)
    refs["bubbles"] = sheet_bubbles(wb)
    refs["construction"] = sheet_construction(wb)
    refs["betas"] = sheet_betas(wb)
    sheet_charts(wb, refs)
    sheet_cover(wb)

    order = ["Cover", "Charts", "CrossAsset", "Claims", "Continuing",
             "Bubbles", "Construction", "RateOilBeta"]
    wb._sheets = [wb[name] for name in order]
    for name in ("Cover", "Charts"):
        wb[name].sheet_properties.tabColor = hx.RED
    wb.active = 0

    hx.set_properties(
        wb, author=AUTHOR, title="Market charts",
        subject="Cross-asset performance, US labour market, rates and construction",
        description="Five charts built from public data: cross-asset rebasing, "
                    "jobless claims, the 10y through equity bubbles, and "
                    "construction spending.",
        keywords="cross-asset; rates; claims; construction",
        category="Markets")

    wb.save(OUT)
    hx.stamp_authorship(OUT, AUTHOR)
    hits = hx.audit_authorship(OUT)
    print("wrote %s (%.1f MB)" % (OUT, os.path.getsize(OUT) / 1e6))
    print("  sheets: %s" % ", ".join(order))
    print("  authorship audit: %s" % ("clean" if not hits else hits))


if __name__ == "__main__":
    main()
