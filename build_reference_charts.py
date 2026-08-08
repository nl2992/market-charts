#!/usr/bin/env python3
"""
Build Rates_Oil_Reference.xlsx — the two charts that cannot be rebuilt from
public data, transcribed from published sell-side research.

INTERNAL REFERENCE. NOT FOR REDISTRIBUTION.

The figures on the PricedIn1y sheet and in the "published" column of the
BetaCompare sheet were read off charts in Citi Research material carrying
"© 2026 Citigroup Inc. No redistribution without Citigroup's written
permission." They are here so the desk can work against them. This workbook
must not go into the public repository, be attached to anything external, or
be reproduced in a client-facing document.

Two sheets:

    PricedIn1y    policy tightening priced over 12 months, by market. There
                  is no free source for this at all - it is OIS-implied
                  pricing - so every figure is transcribed.
    BetaCompare   our own 10y-to-Brent betas against the published ones, for
                  the seven markets where we computed our own. This is the
                  sheet worth looking at: it shows where the two agree and
                  where they do not.

Values are read off charts by eye and are approximate to roughly a bar width.

Run:  python3 data-puller/build_reference_charts.py
"""

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import Reference
from openpyxl.styles import Alignment

import hsbc_xlsx as hx

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "Rates_Oil_Reference.xlsx")

AUTHOR = "Nigel Li"
HDR, R0, C0 = 6, 7, 2

SOURCE = ("Transcribed from Citi Research, © 2026 Citigroup Inc. "
          "No redistribution without Citigroup's written permission.")

# Read off the published bar chart, in basis points priced over 12 months.
PRICED_IN_1Y = [
    ("NZ", 101), ("SE", 76), ("GB", 70), ("CA", 68), ("EU", 62),
    ("JP", 60), ("CH", 48), ("US", 38), ("NO", 18), ("AU", 13),
]

# Published betas, bps per one per cent move in Brent. Ours are computed in
# pull_rate_oil_beta.py on weekly Friday-to-Friday changes since February.
PUBLISHED_BETA = {
    "GB": 0.84, "EU": 0.66, "SE": 0.65, "AU": 0.56, "NO": 0.46,
    "NZ": 0.43, "CA": 0.43, "US": 0.30, "CH": 0.28, "JP": 0.27,
}


def load_our_betas():
    import csv
    path = os.path.join(HERE, "data", "rate_oil_betas.csv")
    with open(path) as fh:
        return {r["code"]: float(r["beta_bps_per_pct_brent"])
                for r in csv.DictReader(fh)}


def grid(ws, first, labels, width=20, first_width=14):
    c = ws.cell(row=HDR, column=1, value=first)
    c.fill, c.font = hx.FILL_RED, hx.F_HEAD
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.column_dimensions["A"].width = first_width
    for i, lab in enumerate(labels):
        c = ws.cell(row=HDR, column=C0 + i, value=lab)
        c.fill, c.font = hx.FILL_RED, hx.F_HEAD
        c.alignment = Alignment(horizontal="right", vertical="center",
                                wrap_text=True)
        ws.column_dimensions[hx.get_column_letter(C0 + i)].width = width
    ws.row_dimensions[HDR].height = 30


def banner(ws, row, width=8):
    """A red strip that says where the numbers came from. Any chart built on
    transcribed figures has to carry its provenance on its face, or it will
    eventually be mistaken for something we measured."""
    for c in range(1, width + 1):
        ws.cell(row=row, column=c).fill = hx.FILL_RED
    cell = ws.cell(row=row, column=1,
                   value="TRANSCRIBED FROM PUBLISHED RESEARCH — "
                         "NOT OUR OWN ESTIMATE — NOT FOR REDISTRIBUTION")
    cell.font = hx.F_HEAD
    ws.row_dimensions[row].height = 18


def sheet_priced(wb, with_chart=True):
    ws = wb.create_sheet("PricedIn1y")
    hx.head(ws, "Priced in over 12 months",
            "Basis points of policy move priced by the market, by currency.",
            width=6)
    banner(ws, 4, width=6)
    grid(ws, "Market", ["bps priced in 1y"], width=20)
    for r, (code, bps) in enumerate(PRICED_IN_1Y, start=R0):
        ws.cell(row=r, column=1, value=code).font = hx.F_LBL
        c = ws.cell(row=r, column=C0, value=bps)
        c.number_format, c.font = "0", hx.F_BODY
    n = len(PRICED_IN_1Y)

    if with_chart:
        vals = Reference(ws, min_col=C0, min_row=R0, max_row=R0 + n - 1)
        cats = Reference(ws, min_col=1, min_row=R0, max_row=R0 + n - 1)
        hx.bar_chart(ws, "E5", "Priced in over 12 months", vals, cats,
                     height=10.0, width=20.0, fmt="0", series_title="bps")

    r = R0 + n + 2
    hx.note(ws, "A%d" % r,
            "There is no free source for this series. It is OIS-implied "
            "policy pricing, which is terminal data. Every figure above was "
            "read off a published bar chart and is approximate to about a bar "
            "width. Government bill yields are not a substitute: they carry "
            "credit and liquidity that OIS does not, so they would move for "
            "reasons unrelated to policy expectations. " + SOURCE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 100
    return ws


def sheet_compare(wb, ours, chart_anchor="G5", with_chart=True):
    ws = wb.create_sheet("BetaCompare")
    hx.head(ws, "10y sensitivity to Brent — ours against the published figures",
            "Basis points per one per cent move in Brent. Ours are computed "
            "on weekly Friday-to-Friday changes since February 2026.", width=8)
    banner(ws, 4, width=8)
    grid(ws, "Market", ["Ours (computed)", "Published (transcribed)",
                        "Difference"], width=19)

    order = sorted(ours, key=lambda c: -ours[c])
    for r, code in enumerate(order, start=R0):
        ws.cell(row=r, column=1, value=code).font = hx.F_LBL
        a = ws.cell(row=r, column=C0, value=ours[code])
        b = ws.cell(row=r, column=C0 + 1, value=PUBLISHED_BETA[code])
        d = ws.cell(row=r, column=C0 + 2,
                    value="=%s%d-%s%d" % (hx.get_column_letter(C0), r,
                                          hx.get_column_letter(C0 + 1), r))
        for c in (a, b):
            c.number_format, c.font = "0.00", hx.F_BODY
        d.number_format, d.font = "+0.00;-0.00", hx.F_LBL
    n = len(order)

    # Three bars per market: ours, theirs, and the gap between them. The gap
    # is the point of the chart — on the US it is the whole argument.
    cats = "'BetaCompare'!$A$%d:$A$%d" % (R0, R0 + n - 1)
    ch = with_chart and hx.grouped_bar_chart(
        ws, chart_anchor, "Ours, published, and the gap",
        [(Reference(ws, min_col=C0 + i, min_row=HDR, max_row=R0 + n - 1),
          colour)
         for i, colour in enumerate((hx.RED, hx.GREY_PALE, hx.NAVY))],
        cats, height=10.0, width=20.0, y_title="bps per % Brent")

    r = R0 + n + 2
    hx.note(ws, "A%d" % r,
            "The two agree on Sweden and the euro area almost exactly, and "
            "agree that Japan is the least sensitive market. They disagree "
            "on the United States: we get roughly twice the published figure, "
            "and across three sampling weekdays and two end dates our US beta "
            "never fell below 0.48. So the claim that the US is among the "
            "least sensitive does not survive our reproduction — Japan is the "
            "outlier, the US is mid-pack. Canada, Switzerland and New Zealand "
            "have no computed figure because no free daily 10y source was "
            "reachable for them. " + SOURCE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 110
    return ws


def sheet_cover(wb):
    ws = wb.create_sheet("Cover", 0)
    hx.head(ws, "Rates and oil — reference charts", "Internal use only",
            width=6)
    banner(ws, 4, width=6)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 86

    r = 6
    for k, v in [
        ("PricedIn1y", "Policy tightening priced over 12 months, by market. "
                       "Every figure transcribed — there is no free source."),
        ("BetaCompare", "Our computed 10y-to-Brent betas against the published "
                        "ones, for the seven markets we could compute."),
    ]:
        hx.label(ws, "A%d" % r, k)
        ws["B%d" % r] = v
        ws["B%d" % r].font = hx.F_BODY
        ws["B%d" % r].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    hx.label(ws, "A%d" % r, "Handling", font=hx.F_H2)
    r += 1
    for line in [
        "This workbook contains figures transcribed from third-party research "
        "that carries an explicit no-redistribution notice. Keep it internal: "
        "it does not go in the public repository, is not attached to anything "
        "external, and is not reproduced in a client-facing document.",
        "Anything we computed ourselves lives in Market_Charts.xlsx, which is "
        "public and carries no third-party figures.",
        "Transcribed values are read off bar charts by eye and are accurate to "
        "roughly a bar width. Do not quote them to two decimal places.",
        "If a number here matters to the argument, source it properly before "
        "it goes in front of anyone.",
    ]:
        r = hx.bullet(ws, r, line)
        ws.merge_cells(start_row=r - 1, start_column=1, end_row=r - 1,
                       end_column=6)
        ws.row_dimensions[r - 1].height = 44

    r += 1
    hx.note(ws, "A%d" % r, SOURCE + "  Built %s."
            % dt.date.today().strftime("%d %B %Y"))
    return ws


def main():
    ours = load_our_betas()
    wb = Workbook()
    wb.remove(wb.active)
    sheet_priced(wb)
    sheet_compare(wb, ours)
    sheet_cover(wb)
    wb._sheets = [wb[n] for n in ("Cover", "PricedIn1y", "BetaCompare")]
    for n in ("Cover", "PricedIn1y", "BetaCompare"):
        wb[n].sheet_properties.tabColor = hx.RED
    wb.active = 0

    hx.set_properties(
        wb, author=AUTHOR,
        title="Rates and oil reference charts — internal, not for redistribution",
        subject="Rates and oil",
        description="Figures transcribed from published research, alongside "
                    "our own computed betas. Internal reference only.",
        keywords="internal; not for redistribution",
        category="Internal reference")

    wb.save(OUT)
    hx.stamp_authorship(OUT, AUTHOR)
    print("wrote %s (%.0f kB)" % (OUT, os.path.getsize(OUT) / 1024))
    print("  INTERNAL ONLY — must not be pushed to the public repository")


if __name__ == "__main__":
    main()
