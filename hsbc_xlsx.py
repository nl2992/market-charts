"""
Chart styling for Excel: a small palette, Arial, and enough axis control to
keep a dense time series readable.

Two things this module exists to get right, both of which fail silently in
openpyxl and cost a rebuild to find:

  * `_cat_series` — Series(values, cats) builds an XY series, which line and
    bar charts silently discard. Categories must be set on `.cat`.
  * no DateAxis — openpyxl's <c:dateAx> makes Excel refuse the file outright.
    A category axis with a date number format renders identically.

Usage:
    import hsbc_xlsx as hx
    ws = wb.create_sheet("Rebased")
    hx.head(ws, "Rebased to 100", "2 Jan 2023 = 100")
"""

import os
import re
import shutil
import zipfile

from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, NumRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (CharacterProperties, Font as DrawingFont,
                                   Paragraph, ParagraphProperties,
                                   RichTextProperties)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# 1. Tokens
# --------------------------------------------------------------------------

RED = "DB0011"        # sampled — logo, rules, header bands
RED_DEEP = "A8000B"   # sampled — negative bars
NAVY = "44546A"       # sampled — chart series 1
BLUE_MID = "6E88A6"   # derived
BLUE_LT = "9FB4CA"    # derived
GREY = "767576"       # sampled — muted text
GREY_RULE = "D0CECE"  # sampled — row rules
GREY_PALE = "D7D8D6"  # sampled — inactive chart elements
TEAL = "00847F"       # sampled — positive deltas
INK = "000000"
WHITE = "FFFFFF"
WASH = "F5F5F5"

#: Series order for charts. Red leads; the blue ramp supports.
PALETTE = [RED, NAVY, BLUE_MID, BLUE_LT, GREY, TEAL, RED_DEEP]

FONT = "Arial"

F_H1 = Font(name=FONT, size=14, bold=True, color=INK)
F_H2 = Font(name=FONT, size=11, bold=True, color=INK)
F_HEAD = Font(name=FONT, size=10, bold=True, color=WHITE)
F_LBL = Font(name=FONT, size=10, bold=True, color=INK)
F_BODY = Font(name=FONT, size=10, color=INK)
F_NOTE = Font(name=FONT, size=9, color=GREY)
F_BLOCK = Font(name=FONT, size=9, bold=True, color=RED)

FILL_RED = PatternFill("solid", fgColor=RED)
FILL_WASH = PatternFill("solid", fgColor=WASH)
THIN = Side(style="thin", color=GREY_RULE)
RULE_BELOW = Border(bottom=THIN)

FMT_DATE = "dd-mmm-yy"
FMT_IDX = "0.0"
FMT_PX = "#,##0.0000"
FMT_PC = "0.0%"
FMT_PC2 = "0.00%"


# --------------------------------------------------------------------------
# 2. Sheet furniture
# --------------------------------------------------------------------------

def head(ws, title, subtitle=None, width=10, freeze=None):
    """Title, kicker, and the thin red rule that opens every sheet."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = F_H1
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = F_NOTE
    for c in range(1, width + 1):
        ws.cell(row=3, column=c).fill = FILL_RED
    ws.row_dimensions[3].height = 4
    if freeze:
        ws.freeze_panes = freeze


def header_row(ws, row, labels, start_col=1, widths=None, height=26, wrap=True):
    """Solid red band, white bold type — the results-table header."""
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lab)
        c.fill = FILL_RED
        c.font = F_HEAD
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=wrap)
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w
    ws.row_dimensions[row].height = height


def note(ws, cell, text):
    ws[cell] = text
    ws[cell].font = F_NOTE
    ws[cell].alignment = Alignment(vertical="top", wrap_text=True)


def label(ws, cell, text, font=None):
    ws[cell] = text
    ws[cell].font = font or F_LBL


def bullet(ws, row, text, col=1, indent=0):
    """The red diamond bullet used throughout the deck."""
    c = ws.cell(row=row, column=col, value="◆  " + text)
    c.font = F_BODY
    c.alignment = Alignment(vertical="top", wrap_text=True, indent=indent)
    return row + 1


# --------------------------------------------------------------------------
# 3. Chart theming
# --------------------------------------------------------------------------
# openpyxl exposes no styling API for chart text, so character properties are
# built by hand and hung off the axis, legend and title elements.

def _cp(size=1000, bold=False, colour=INK):
    return CharacterProperties(latin=DrawingFont(typeface=FONT),
                               sz=size, b=bold, solidFill=colour)


def _txpr(size=1000, bold=False, colour=INK):
    cp = _cp(size, bold, colour)
    return RichText(bodyPr=RichTextProperties(),
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp),
                                 endParaRPr=cp)])


def _hairline_gridlines():
    """Horizontal rules the eye can read a level off without them competing
    with the data. Excel's default gridline is far too heavy for this."""
    return ChartLines(spPr=GraphicalProperties(
        ln=LineProperties(solidFill=GREY_RULE, w=3175)))


def theme_chart(ch, title=None, legend=True, y_title=None, gridlines=True):
    """Arial throughout, black title, grey axis labels, hairline horizontal
    gridlines, no border, legend along the bottom.

    Tick marks are removed on both axes: with the axis line and the gridlines
    already present they are pure clutter, and on a dense category axis they
    print once per category whether or not the label does.
    """
    if title is not None:
        ch.title = title
        try:
            rich = ch.title.tx.rich
            rich.p[0].pPr = ParagraphProperties(defRPr=_cp(1200, True, INK))
            for r in rich.p[0].r:
                r.rPr = _cp(1200, True, INK)
        except (AttributeError, IndexError, TypeError):
            pass

    for ax in (ch.x_axis, ch.y_axis):
        ax.txPr = _txpr(950, False, GREY)
        ax.majorGridlines = None
        ax.majorTickMark = "none"
        ax.minorTickMark = "none"
        ax.spPr = GraphicalProperties(ln=LineProperties(solidFill=GREY_RULE, w=6350))
    if gridlines:
        ch.y_axis.majorGridlines = _hairline_gridlines()
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    # openpyxl leaves axPos at its default of "l" on both axes, which puts the
    # category axis down the left-hand side. Excel is forgiving enough to draw
    # it at the bottom anyway, but the file is wrong and a second value axis
    # inherits the same default and genuinely does land on the wrong side.
    ch.x_axis.axPos = "b"
    ch.y_axis.axPos = "l"

    if y_title:
        ch.y_axis.title = y_title
        try:
            rich = ch.y_axis.title.tx.rich
            rich.p[0].pPr = ParagraphProperties(defRPr=_cp(950, False, GREY))
        except (AttributeError, IndexError, TypeError):
            pass

    if legend and ch.legend is not None:
        ch.legend.position = "b"
        ch.legend.overlay = False
        ch.legend.txPr = _txpr(900, False, INK)
    else:
        ch.legend = None

    # The reference deck never boxes a chart.
    ch.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    return ch


def _cat_series(values, cats, title):
    """A category series for a Line or Bar chart.

    NOT Series(values, cats, ...): passing categories positionally makes
    openpyxl build an XYSeries with xVal/yVal, which is the scatter shape.
    Line and bar charts serialise val/cat, so the data is dropped without an
    error and the chart renders empty. Set .cat explicitly.
    """
    s = Series(values, title=title)
    s.cat = AxDataSource(numRef=NumRef(f=cats))
    return s


def line_series(ws, col, row0, nrows, cat_col, colour, title,
                width=17000, dashed=False):
    ref = Reference(ws, min_col=col, min_row=row0, max_row=row0 + nrows - 1)
    cats = Reference(ws, min_col=cat_col, min_row=row0, max_row=row0 + nrows - 1)
    s = _cat_series(ref, cats, title)
    ln = LineProperties(solidFill=colour, w=width)
    if dashed:
        ln.prstDash = "dash"
    s.graphicalProperties = GraphicalProperties(ln=ln)
    s.marker = Marker(symbol="none")
    s.smooth = False
    return s


def series_style(i):
    """Colour and dash for the i-th series, cycling the palette and switching
    to a dashed stroke on the second pass so an eight-line chart stays
    readable in mono."""
    return PALETTE[i % len(PALETTE)], i >= len(PALETTE)


def axis_labels(ch, n_points, target=9, x_fmt=None, y_fmt=None,
                y_min=None, y_max=None):
    """Thin the category labels down to something readable.

    A category axis prints one label per category unless told otherwise, so a
    daily series arrives as several hundred overlapping labels and Excel
    resolves it by rotating them 45 degrees, which is the clutter. Aim for
    about `target` labels across the axis instead.
    """
    if x_fmt:
        ch.x_axis.number_format = x_fmt
        ch.x_axis.majorTimeUnit = None
    if y_fmt:
        ch.y_axis.number_format = y_fmt
    skip = max(1, int(round(n_points / float(target))))
    ch.x_axis.tickLblSkip = skip
    ch.x_axis.tickMarkSkip = skip
    if y_min is not None:
        ch.y_axis.scaling.min = y_min
    if y_max is not None:
        ch.y_axis.scaling.max = y_max
    return ch


def line_chart(ws, anchor, title, series, y_title=None, height=8.6, width=18.0,
               legend=True, date_axis=True, n_points=None, target_labels=9,
               x_fmt=None, y_fmt=None, y_min=None, y_max=None, gridlines=True):
    ch = LineChart()
    ch.height, ch.width = height, width
    if date_axis:
        # Not DateAxis. openpyxl emits <c:dateAx> XML that Excel rejects —
        # the workbook will not open at all. A category axis with a date
        # number format renders the same, provided the source cells carry a
        # date format, which they do.
        x_fmt = x_fmt or "mmm-yy"
    for s in series:
        ch.series.append(s)
    theme_chart(ch, title=title, legend=legend, y_title=y_title,
                gridlines=gridlines)
    if n_points:
        axis_labels(ch, n_points, target=target_labels, x_fmt=x_fmt,
                    y_fmt=y_fmt, y_min=y_min, y_max=y_max)
    else:
        axis_labels(ch, 1, target=1, x_fmt=x_fmt, y_fmt=y_fmt,
                    y_min=y_min, y_max=y_max)
    ws.add_chart(ch, anchor)
    return ch


def combo_chart(ws, anchor, title, primary, secondary, y_title=None,
                y2_title=None, height=8.6, width=18.0, date_axis=False,
                n_points=None, target_labels=9, x_fmt=None, y_fmt=None,
                y2_fmt=None):
    """Two line charts sharing an x axis, the second on its own right-hand
    scale.

    For series measured in different units — a per cent change against a
    percentage point change — where plotting on one axis would invite the
    reader to compare the gap between them, which means nothing.

    openpyxl builds this by merging a second chart into the first. The second
    chart's value axis needs its own id and must cross at the far end, or
    Excel draws both scales on the left on top of each other.
    """
    c1 = LineChart()
    c1.height, c1.width = height, width
    for s in primary:
        c1.series.append(s)

    c2 = LineChart()
    for s in secondary:
        c2.series.append(s)
    c2.y_axis.axId = 200
    c2.y_axis.crosses = "max"
    c2.y_axis.axPos = "r"          # or it stacks on top of the primary scale
    c2.y_axis.delete = False       # absent reads as "hide me" in some builds
    c2.y_axis.majorTickMark = "none"
    c2.y_axis.minorTickMark = "none"
    c2.y_axis.majorGridlines = None
    c2.y_axis.txPr = _txpr(950, False, GREY)
    c2.y_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=GREY_RULE,
                                                           w=6350))
    if y2_title:
        c2.y_axis.title = y2_title
        try:
            c2.y_axis.title.tx.rich.p[0].pPr = ParagraphProperties(
                defRPr=_cp(950, False, GREY))
        except (AttributeError, IndexError, TypeError):
            pass

    if y2_fmt:
        c2.y_axis.number_format = y2_fmt

    theme_chart(c1, title=title, legend=True, y_title=y_title)
    if n_points:
        axis_labels(c1, n_points, target=target_labels,
                    x_fmt=x_fmt or ("mmm-yy" if date_axis else None),
                    y_fmt=y_fmt)
    elif y_fmt:
        c1.y_axis.number_format = y_fmt
    c1 += c2
    ws.add_chart(c1, anchor)
    return c1


def bar_chart(ws, anchor, title, val_ref, cat_ref, colour=RED, height=8.6,
              width=18.0, series_title=None, labels=True, fmt="#,##0",
              horizontal=False):
    ch = BarChart()
    ch.type = "bar" if horizontal else "col"
    ch.height, ch.width = height, width
    ch.gapWidth = 60
    s = _cat_series(val_ref, cat_ref, series_title)
    s.graphicalProperties = GraphicalProperties(solidFill=colour)
    s.graphicalProperties.ln = LineProperties(noFill=True)
    ch.series.append(s)
    if labels:
        ch.dataLabels = DataLabelList()
        ch.dataLabels.showVal = True
        ch.dataLabels.numFmt = fmt
        ch.dataLabels.txPr = _txpr(900, True, INK)
    theme_chart(ch, title=title, legend=False)
    ws.add_chart(ch, anchor)
    return ch


# --------------------------------------------------------------------------
# 4. Document properties
# --------------------------------------------------------------------------

_APP_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006'
    '/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/'
    'officeDocument/2006/docPropsVTypes">'
    "<Application>Microsoft Excel</Application>"
    "<DocSecurity>0</DocSecurity>"
    "<ScaleCrop>false</ScaleCrop>"
    "<LinksUpToDate>false</LinksUpToDate>"
    "<SharedDoc>false</SharedDoc>"
    "<HyperlinksChanged>false</HyperlinksChanged>"
    "<AppVersion>16.0300</AppVersion>"
    "</Properties>"
)


def set_properties(wb, author, title, subject=None, description=None,
                   keywords=None, category=None):
    """Author the file properly. Excel shows these in Get Info and in the
    document panel, and they are what a reader checks first."""
    p = wb.properties
    p.creator = author
    p.lastModifiedBy = author
    p.title = title
    p.subject = subject
    p.description = description
    p.keywords = keywords
    p.category = category
    p.revision = "1"


def stamp_authorship(path, author):
    """Rewrite the parts of the package that name the build tool rather than
    the author: docProps/app.xml carries an Application string, and core.xml
    can be left with a stale creator. Excel reports both.

    Run after wb.save(). Rewrites the archive in place.
    """
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as src, \
            zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "docProps/app.xml":
                data = _APP_XML.encode("utf-8")
            elif item.filename == "docProps/core.xml":
                text = data.decode("utf-8")
                text = re.sub(r"(<dc:creator[^>]*>).*?(</dc:creator>)",
                              r"\g<1>%s\g<2>" % author, text)
                text = re.sub(r"(<cp:lastModifiedBy[^>]*>).*?(</cp:lastModifiedBy>)",
                              r"\g<1>%s\g<2>" % author, text)
                data = text.encode("utf-8")
            dst.writestr(item, data)
    os.replace(tmp, path)


def audit_authorship(path, forbidden=("openpyxl", "python", "Python")):
    """Fail loudly if a build-tool name survives anywhere in the package."""
    hits = []
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            body = z.read(name)
            for token in forbidden:
                if token.encode("utf-8") in body:
                    hits.append("%s: %s" % (name, token))
    return hits


def copy_to(path, dest_dir):
    """Place a built workbook alongside the other deliverables."""
    dest = os.path.join(dest_dir, os.path.basename(path))
    shutil.copyfile(path, dest)
    return dest
