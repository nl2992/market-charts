#!/usr/bin/env python3
"""
Pull the four series sets behind the underweight-fixed-income exhibits.

The argument these support is not the inflation one. It is that the labour
market is not deteriorating, that yields have historically risen through
equity bubbles rather than fallen, and that a construction capex boom is
bidding for the same savings duration needs. All four cut the same way:
you are not paid to own duration here.

Everything comes from FRED, which fronts BLS and Census. Free, no key,
and every series is one a PM can pull up and check.

    claims_weekly.csv     initial and continuing claims, weekly
    urate_monthly.csv     unemployment rate, monthly
    bubbles.csv           10y Treasury path through six equity episodes
    construction.csv      nonresidential construction, deflated

Run:  python3 data-puller/pull_duration_case.py
"""

import csv
import datetime as dt
import os
import time
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "data")

# FRED refuses a browser user agent on this endpoint; the default is fine.
FRED = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=%s&cosd=%s&coed=%s"

# The Census C30 seasonally adjusted history. Needed because the data centre
# line is not on FRED at all.
CENSUS_VIP = "https://www.census.gov/construction/c30/xls/privsatime.xlsx"

# Equity episodes, as used in the sell-side version of this chart. The last
# one is the current cycle and is still running.
EPISODES = [
    ("1964-02-29", "1966-09-30", "1964-66"),
    ("1997-07-31", "1998-06-30", "1997-98"),
    ("1998-11-30", "2002-09-30", "1998-2002 dot-com"),
    ("2020-10-31", "2022-09-30", "2020-22"),
    ("2023-06-30", "2025-04-08", "2023-25"),
    ("2025-06-26", "2026-07-23", "2025-26 current"),
]

# The reference week for the payroll and household surveys is the week
# containing the 12th. Weeks end on a Saturday, so exactly one week-ending
# date falls in the 12th-to-18th window each month.
REF_LO, REF_HI = 12, 18


def fred(series_id, start, end, tries=5):
    """One FRED series as {date: value}.

    Retries hard and paces between calls. FRED throttles a burst by returning
    404 rather than 429, so a missing series and a rate limit look identical —
    hence the backoff before believing the first failure.
    """
    url = FRED % (series_id, start, end)
    for i in range(tries):
        try:
            with urllib.request.urlopen(url, timeout=60) as r:
                body = r.read().decode("utf-8")
            break
        except Exception:
            if i == tries - 1:
                raise
            time.sleep(5 * (i + 1))
    lines = body.splitlines()
    if not lines or not lines[0].lower().startswith("observation_date"):
        raise RuntimeError("%s did not return a CSV" % series_id)
    out = {}
    for row in csv.reader(lines[1:]):
        if len(row) < 2 or row[1] in (".", ""):
            continue
        out[dt.date.fromisoformat(row[0])] = float(row[1])
    if not out:
        raise RuntimeError("%s returned no observations" % series_id)
    print("  %-14s %5d obs  %s to %s"
          % (series_id, len(out), min(out), max(out)))
    time.sleep(4)
    return out


def write(name, header, rows):
    path = os.path.join(OUT_DIR, name)
    with open(path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)
    print("wrote %s — %d rows" % (path, len(rows)))


def claims_by_week(icsa):
    """Initial claims laid out by week of year, one column per calendar year.

    Plotting the seasonally adjusted series this way is the point: if the
    adjustment were doing its job the years would sit on top of one another,
    and where the current year runs below the others it is running below its
    own seasonal, not just below last year.
    """
    years = [2023, 2024, 2025, 2026]
    grid = {y: {} for y in years}
    for d, v in icsa.items():
        if d.year in grid:
            grid[d.year][d.isocalendar().week] = v
    rows = []
    for wk in range(1, 54):
        row = [wk] + [grid[y].get(wk, "") for y in years]
        if any(c != "" for c in row[1:]):
            rows.append(row)
    return ["week"] + [str(y) for y in years], rows


def reference_week(ccsa, year, month):
    """The continuing-claims print for the survey reference week."""
    for d, v in ccsa.items():
        if d.year == year and d.month == month and REF_LO <= d.day <= REF_HI:
            return v
    return None


def claims_vs_urate(ccsa, unrate):
    """Continuing claims and the unemployment rate against the same base
    month, for this cycle and for the 2024 analogue."""
    cohorts = [(2025, "2026"), (2023, "2024")]
    months = [(11, 0), (12, 1), (1, 2), (2, 3), (3, 4), (4, 5), (5, 6),
              (6, 7), (7, 8)]
    series = {}
    for base_year, tag in cohorts:
        cc0 = ur0 = None
        cc, ur = [], []
        for month, step in months:
            y = base_year if month >= 11 else base_year + 1
            c = reference_week(ccsa, y, month)
            u = unrate.get(dt.date(y, month, 1))
            if step == 0:
                cc0, ur0 = c, u
            cc.append("" if c is None or cc0 is None else round(c / cc0 - 1, 6))
            ur.append("" if u is None or ur0 is None else round(u - ur0, 4))
        series["cc_" + tag], series["ur_" + tag] = cc, ur

    labels = ["Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul"]
    header = ["month", "cc_2026", "cc_2024", "ur_2026", "ur_2024"]
    rows = [[labels[i], series["cc_2026"][i], series["cc_2024"][i],
             series["ur_2026"][i], series["ur_2024"][i]]
            for i in range(len(labels))]
    return header, rows


def bubble_paths(dgs10):
    """Change in the 10y from the start of each episode, by trading day."""
    dates = sorted(dgs10)
    cols, longest = [], 0
    for start, end, label in EPISODES:
        s, e = dt.date.fromisoformat(start), dt.date.fromisoformat(end)
        path = [dgs10[d] for d in dates if s <= d <= e]
        if not path:
            raise RuntimeError("no 10y data for episode %s" % label)
        cols.append((label, [round(v - path[0], 4) for v in path]))
        longest = max(longest, len(path))
        print("  %-20s %4d trading days, %+.2fpp end to end"
              % (label, len(path), path[-1] - path[0]))
    header = ["day"] + [c[0] for c in cols]
    rows = [[i] + [(c[1][i] if i < len(c[1]) else "") for c in cols]
            for i in range(longest)]
    return header, rows


def census_vip():
    """Private construction spending by type, from the Census C30 release.

    FRED carries 144 series from this release and none of them is the data
    centre line — the whole point of the chart — so this goes to the source
    file. Column 10 of the seasonally adjusted history is 'Data center',
    published monthly back to January 2014.
    """
    from openpyxl import load_workbook

    path = os.path.join(OUT_DIR, "_census_privsatime.xlsx")
    # Census 403s the default Python user agent; FRED 403s a browser one.
    req = urllib.request.Request(CENSUS_VIP, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=90) as r:
        body = r.read()
    with open(path, "wb") as fh:
        fh.write(body)

    ws = load_workbook(path)["Private SA"]
    header = [str(c or "").replace("_x000D_", " ").strip()
              for c in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
    for want in ("Nonresidential", "Data center"):
        if want not in header:
            raise RuntimeError("Census layout changed: no '%s' column" % want)
    i_nonres, i_dc = header.index("Nonresidential"), header.index("Data center")

    out = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        stamp = row[0]
        if not stamp or not isinstance(stamp, str):
            continue
        try:                       # 'Jun-26p' and 'May-26r' carry a revision flag
            d = dt.datetime.strptime(stamp.rstrip("pr"), "%b-%y").date()
        except ValueError:
            continue
        nonres, dc = row[i_nonres], row[i_dc]
        if isinstance(nonres, (int, float)) and isinstance(dc, (int, float)):
            out[d] = (float(nonres), float(dc))
    os.remove(path)
    print("  %-14s %5d obs  %s to %s  (Census C30)"
          % ("private VIP", len(out), min(out), max(out)))
    return out


def construction(vip, ppi):
    """Private nonresidential construction split into the data centre boom and
    everything else, as a change since January 2024 in $bn.

    Both bases are written. Nominal is the like-for-like reproduction of how
    the sell-side draws this. Real is the stronger version of the argument:
    construction costs rose through the window, so the nominal line flatters
    everything that is not the boom. Chart nominal, quote real.
    """
    base = dt.date(2024, 1, 1)
    months = sorted(d for d in vip if d >= base and d in ppi)
    if not months or months[0] != base:
        raise RuntimeError("January 2024 base month is missing")
    p0 = ppi[base]
    t0_n, d0_n = vip[base]

    def deflated(d):
        nonres, dc = vip[d]
        f = ppi[d] / p0
        return nonres / f, dc / f

    t0_r, d0_r = deflated(base)
    rows = []
    for d in months:
        tn, cn = vip[d]
        tr, cr = deflated(d)
        rows.append([
            d.isoformat(),
            round((tn - t0_n) / 1000.0, 3),
            round((cn - d0_n) / 1000.0, 3),
            round(((tn - cn) - (t0_n - d0_n)) / 1000.0, 3),
            round((tr - t0_r) / 1000.0, 3),
            round((cr - d0_r) / 1000.0, 3),
            round(((tr - cr) - (t0_r - d0_r)) / 1000.0, 3),
        ])
    last = rows[-1]
    print("  %d months, Jan-24 to %s" % (len(rows), months[-1]))
    print("    nominal $bn: total %+.1f, data centre %+.1f, ex data centre %+.1f"
          % (last[1], last[2], last[3]))
    print("    Jan-24 prices: total %+.1f, data centre %+.1f, ex data centre %+.1f"
          % (last[4], last[5], last[6]))
    return (["date", "total", "data_center", "ex_data_center",
             "total_real", "data_center_real", "ex_data_center_real"], rows)


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    today = dt.date.today().isoformat()

    print("claims:")
    icsa = fred("ICSA", "2022-12-01", today)
    ccsa = fred("CCSA", "2023-10-01", today)
    unrate = fred("UNRATE", "2023-10-01", today)

    header, rows = claims_by_week(icsa)
    write("claims_weekly.csv", header, rows)
    header, rows = claims_vs_urate(ccsa, unrate)
    write("claims_urate.csv", header, rows)

    print("10y through equity episodes:")
    dgs10 = fred("DGS10", "1962-01-01", today)
    header, rows = bubble_paths(dgs10)
    write("bubbles.csv", header, rows)

    print("construction:")
    vip = census_vip()
    ppi = fred("WPUSI012011", "2023-12-01", today)
    header, rows = construction(vip, ppi)
    write("construction.csv", header, rows)


if __name__ == "__main__":
    main()
